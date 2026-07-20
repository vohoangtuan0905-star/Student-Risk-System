import mysql.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

conn = mysql.connector.connect(
    host='127.0.0.1', port=3306,
    user='root', password='',
    database='student_risk_db', charset='utf8mb4'
)
cur = conn.cursor(dictionary=True)

BLUE_DARK  = 'FF1E3A5F'
GREEN_DARK = 'FF1A5C38'
GREEN_LIGHT= 'FFD4EDDA'
YELLOW     = 'FFFFF3CD'
WHITE      = 'FFFFFFFF'
GRAY       = 'FFF2F2F2'

def header_cell(ws, row, col, value, bg=BLUE_DARK):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
    c.fill = PatternFill(fill_type='solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FF999999')
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_cell(ws, row, col, value, bg=WHITE, align='left', num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=10, name='Calibri')
    c.fill = PatternFill(fill_type='solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    thin = Side(style='thin', color='FFCCCCCC')
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if num_fmt:
        c.number_format = num_fmt

# ====================================================
# FILE 1: mau_import_thongtin_sv.xlsx (khong doi)
# ====================================================
cur.execute("""
    SELECT s.student_code, s.full_name,
           DATE_FORMAT(s.date_of_birth, '%d/%m/%Y') AS date_of_birth,
           CASE s.gender WHEN 'Male' THEN 'Nam' WHEN 'Female' THEN 'Nu' ELSE 'Khac' END AS gender,
           s.email, s.phone,
           COALESCE(s.address, 'TP. Ho Chi Minh') AS address,
           c.class_code, c.class_name, d.department_name,
           s.enrollment_year
    FROM students s
    JOIN classes c ON c.id = s.class_id
    JOIN departments d ON d.id = s.department_id
    WHERE c.class_code LIKE 'D22_%'
    ORDER BY d.department_name, c.class_code, s.student_code
    LIMIT 30
""")
sv_rows = cur.fetchall()

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'Import_ThongTin_SV'
ws1.sheet_view.showGridLines = False

ws1.merge_cells('A1:K1')
t = ws1['A1']
t.value = 'MAU IMPORT THONG TIN SINH VIEN - HK2 NAM HOC 2025-2026'
t.font = Font(bold=True, size=13, color='FFFFFFFF', name='Calibri')
t.fill = PatternFill(fill_type='solid', fgColor=BLUE_DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:K2')
n = ws1['A2']
n.value = 'Huong dan: Giu nguyen ten cot dong 3. Ma sinh vien (student_code) va Ma lop (class_code) la bat buoc. Gioi tinh: Nam/Nu/Khac. Ngay sinh: DD/MM/YYYY.'
n.font = Font(italic=True, size=9, color='FF7D4E00', name='Calibri')
n.fill = PatternFill(fill_type='solid', fgColor=YELLOW)
n.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws1.row_dimensions[2].height = 22

COLS_INFO = [
    ('Ma sinh vien\n(student_code)', 16),
    ('Ho va ten\n(full_name)', 24),
    ('Ngay sinh\n(date_of_birth)', 14),
    ('Gioi tinh\n(gender)', 10),
    ('Email\n(email)', 28),
    ('So dien thoai\n(phone)', 14),
    ('Dia chi\n(address)', 28),
    ('Ma lop\n(class_code)', 14),
    ('Ten lop\n(class_name)', 22),
    ('Khoa\n(department_name)', 22),
    ('Nam nhap hoc\n(enrollment_year)', 14),
]
for ci, (col_name, col_w) in enumerate(COLS_INFO, 1):
    header_cell(ws1, 3, ci, col_name)
    ws1.column_dimensions[get_column_letter(ci)].width = col_w
ws1.row_dimensions[3].height = 34

for ri, row in enumerate(sv_rows, 4):
    bg = WHITE if ri % 2 == 0 else GRAY
    vals = [
        row['student_code'], row['full_name'], row['date_of_birth'],
        row['gender'], row['email'], row['phone'], row['address'],
        row['class_code'], row['class_name'], row['department_name'],
        row['enrollment_year']
    ]
    for ci, v in enumerate(vals, 1):
        data_cell(ws1, ri, ci, v, bg=bg, align='center' if ci in [1,3,4,8,11] else 'left')
    ws1.row_dimensions[ri].height = 18

ws1.freeze_panes = 'A4'
wb1.save(r'd:\Student-Risk-System\frontend\public\mau_import_thongtin_sv.xlsx')
print('File 1 OK')

# ====================================================
# FILE 2: mau_import_ketqua_hoctap.xlsx
# KHONG co: warning_level, actual_status (he thong tu tinh)
# Chi giu: thong tin hoc tap thuan tuy
# ====================================================
cur.execute("""
    SELECT s.student_code, s.full_name,
           c.class_code,
           2 AS hoc_ky,
           '2025-2026' AS nam_hoc,
           ROUND(sar.gpa, 2) AS gpa,
           sar.absences,
           sar.tuition_debt,
           sar.scholarship,
           sar.failed_subjects,
           sar.credits_enrolled,
           sar.credits_passed
    FROM student_academic_records sar
    JOIN students s ON s.id = sar.student_id
    JOIN classes c ON c.id = s.class_id
    WHERE sar.semester_id = 4
      AND c.class_code LIKE 'D22_%'
    ORDER BY c.class_code, s.student_code
    LIMIT 50
""")
kq_rows = cur.fetchall()

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Import_KetQua_HocTap'
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:L1')
t2 = ws2['A1']
t2.value = 'MAU IMPORT KET QUA HOC TAP - HK2 NAM HOC 2025-2026 (Ket thuc 15/06/2026)'
t2.font = Font(bold=True, size=13, color='FFFFFFFF', name='Calibri')
t2.fill = PatternFill(fill_type='solid', fgColor=GREEN_DARK)
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:L2')
n2 = ws2['A2']
n2.value = ('Huong dan: Cot Hoc ky (1/2/3) va Nam hoc (VD: 2025-2026) bat buoc. '
            'GPA: 0.00-4.00 | No hoc phi: 0=Khong, 1=Co | Hoc bong: 0=Khong, 1=Co | '
            'He thong tu dong tinh muc rui ro sau khi import.')
n2.font = Font(italic=True, size=9, color='FF1A5C38', name='Calibri')
n2.fill = PatternFill(fill_type='solid', fgColor=GREEN_LIGHT)
n2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[2].height = 22

# Header - chi co 12 cot, KHONG co warning_level va actual_status
COLS_KQ = [
    ('Ma sinh vien\n(student_code)', 16),
    ('Ho va ten\n(full_name)', 22),
    ('Ma lop\n(class_code)', 14),
    ('Hoc ky\n(semester_no)', 10),
    ('Nam hoc\n(academic_year)', 12),
    ('GPA\n(gpa)', 8),
    ('So buoi vang\n(absences)', 12),
    ('No hoc phi\n(tuition_debt)\n0=Khong / 1=Co', 14),
    ('Hoc bong\n(scholarship)\n0=Khong / 1=Co', 14),
    ('Mon thi lai\n(failed_subjects)', 12),
    ('TC dang ky\n(credits_enrolled)', 12),
    ('TC da qua\n(credits_passed)', 12),
]
for ci, (col_name, col_w) in enumerate(COLS_KQ, 1):
    header_cell(ws2, 3, ci, col_name, bg=GREEN_DARK)
    ws2.column_dimensions[get_column_letter(ci)].width = col_w
ws2.row_dimensions[3].height = 40

for ri, row in enumerate(kq_rows, 4):
    bg = WHITE if ri % 2 == 0 else GRAY
    vals = [
        row['student_code'], row['full_name'], row['class_code'],
        row['hoc_ky'], row['nam_hoc'],
        row['gpa'], row['absences'],
        row['tuition_debt'], row['scholarship'],
        row['failed_subjects'], row['credits_enrolled'], row['credits_passed']
    ]
    for ci, v in enumerate(vals, 1):
        data_cell(ws2, ri, ci, v, bg=bg,
                  align='center' if ci in [1,3,4,5,8,9] else 'left',
                  num_fmt='0.00' if ci==6 else None)
    ws2.row_dimensions[ri].height = 18

ws2.freeze_panes = 'A4'
wb2.save(r'd:\Student-Risk-System\frontend\public\mau_import_ketqua_hoctap.xlsx')
print('File 2 OK')

cur.close()
conn.close()
print('HOAN THANH - 2 file da cap nhat!')
